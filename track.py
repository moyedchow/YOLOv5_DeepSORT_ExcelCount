# limit the number of cpus used by high performance libraries
import os
from re import X
from tarfile import REGULAR_TYPES
from tkinter import Y

from numpy import var
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

import sys
sys.path.insert(0, './yolov5')

import argparse
import os
import platform
import shutil
import time
from pathlib import Path
import cv2
import torch
import torch.backends.cudnn as cudnn
import xlsxwriter
import datetime
import openpyxl

from yolov5.models.experimental import attempt_load
from yolov5.utils.downloads import attempt_download
from yolov5.models.common import DetectMultiBackend
from yolov5.utils.datasets import LoadImages, LoadStreams
from yolov5.utils.general import (LOGGER, check_img_size, non_max_suppression, scale_coords, 
                                  check_imshow, xyxy2xywh, increment_path)
from yolov5.utils.torch_utils import select_device, time_sync
from yolov5.utils.plots import Annotator, colors
from deep_sort.utils.parser import get_config
from deep_sort.deep_sort import DeepSort

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # yolov5 deepsort root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative
count = 0
data = []
count_reg = 0
data_reg = []
count_tiff = 0
data_tiff = []

def detect(opt):
    out, source, yolo_model, deep_sort_model, show_vid, save_vid, save_txt, imgsz, evaluate, half, project, name, exist_ok= \
        opt.output, opt.source, opt.yolo_model, opt.deep_sort_model, opt.show_vid, opt.save_vid, \
        opt.save_txt, opt.imgsz, opt.evaluate, opt.half, opt.project, opt.name, opt.exist_ok
    webcam = source == '0' or source.startswith(
        'rtsp') or source.startswith('http') or source.endswith('.txt')

    # initialize deepsort
    cfg = get_config()
    cfg.merge_from_file(opt.config_deepsort)
    deepsort = DeepSort(deep_sort_model,
                        max_dist=cfg.DEEPSORT.MAX_DIST,
                        max_iou_distance=cfg.DEEPSORT.MAX_IOU_DISTANCE,
                        max_age=cfg.DEEPSORT.MAX_AGE, n_init=cfg.DEEPSORT.N_INIT, nn_budget=cfg.DEEPSORT.NN_BUDGET,
                        use_cuda=True)

    # Initialize
    device = select_device(opt.device)
    half &= device.type != 'cpu'  # half precision only supported on CUDA

    # The MOT16 evaluation runs multiple inference streams in parallel, each one writing to
    # its own .txt file. Hence, in that case, the output folder is not restored
    if not evaluate:
        if os.path.exists(out):
            pass
            shutil.rmtree(out)  # delete output folder
        os.makedirs(out)  # make new output folder

    # Directories
    save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
    save_dir.mkdir(parents=True, exist_ok=True)  # make dir

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(yolo_model, device=device, dnn=opt.dnn)
    stride, names, pt, jit, _ = model.stride, model.names, model.pt, model.jit, model.onnx
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Half
    half &= pt and device.type != 'cpu'  # half precision only supported by PyTorch on CUDA
    if pt:
        model.model.half() if half else model.model.float()

    # Set Dataloader
    vid_path, vid_writer = None, None
    # Check if environment supports image displays
    if show_vid:
        show_vid = check_imshow()

    # Dataloader
    if webcam:
        show_vid = check_imshow()
        cudnn.benchmark = True  # set True to speed up constant image size inference
        dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt and not jit)
        bs = len(dataset)  # batch_size
    else:
        dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)
        bs = 1  # batch_size
    vid_path, vid_writer = [None] * bs, [None] * bs

    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names

    # extract what is in between the last '/' and last '.'
    txt_file_name = source.split('/')[-1].split('.')[0]
    txt_path = str(Path(save_dir)) + '/' + txt_file_name + '.txt'

    if pt and device.type != 'cpu':
        model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
    dt, seen = [0.0, 0.0, 0.0, 0.0], 0
    for frame_idx, (path, img, im0s, vid_cap, s) in enumerate(dataset):
        t1 = time_sync()
        img = torch.from_numpy(img).to(device)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)
        t2 = time_sync()
        dt[0] += t2 - t1

        # Inference
        visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if opt.visualize else False
        pred = model(img, augment=opt.augment, visualize=visualize)
        t3 = time_sync()
        dt[1] += t3 - t2

        # Apply NMS
        pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, opt.classes, opt.agnostic_nms, max_det=opt.max_det)
        dt[2] += time_sync() - t3

        # Process detections
        for i, det in enumerate(pred):  # detections per image
            seen += 1
            if webcam:  # batch_size >= 1
                p, im0, _ = path[i], im0s[i].copy(), dataset.count
                s += f'{i}: '
            else:
                p, im0, _ = path, im0s.copy(), getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg, vid.mp4, ...
            s += '%gx%g ' % img.shape[2:]  # print string

            annotator = Annotator(im0, line_width=2, pil=not ascii)
            w, h = im0.shape[1],im0.shape[0]
            if det is not None and len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_coords(
                    img.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                xywhs = xyxy2xywh(det[:, 0:4])
                confs = det[:, 4]
                clss = det[:, 5]

                # pass detections to deepsort
                t4 = time_sync()
                outputs = deepsort.update(xywhs.cpu(), confs.cpu(), clss.cpu(), im0)
                t5 = time_sync()
                dt[3] += t5 - t4

                # draw boxes for visualization
                if len(outputs) > 0:
                    for j, (output, conf) in enumerate(zip(outputs, confs)):

                        bboxes = output[0:4]
                        id = output[4]
                        cls = output[5]
                        #count                        
                        count_obj(bboxes,w,h,id,cls)
                        c = int(cls)  # integer class
                        label = f' {names[c]} ' #label = f'{id} {names[c]} {conf:.2f}'
                        annotator.box_label(bboxes, label, color=colors(c, True))

                        if save_txt:
                            # to MOT format
                            bbox_left = output[0]
                            bbox_top = output[1]
                            bbox_w = output[2] - output[0]
                            bbox_h = output[3] - output[1]
                            # Write MOT compliant results to file
                            with open(txt_path, 'a') as f:
                                f.write(('%g ' * 10 + '\n') % (frame_idx + 1, id, bbox_left,  # MOT format
                                                               bbox_top, bbox_w, bbox_h, -1, -1, -1, -1))

                LOGGER.info(f'{s}Done. YOLO:({t3 - t2:.3f}s), DeepSort:({t5 - t4:.3f}s)')

            else:
                deepsort.increment_ages()
                LOGGER.info('No detections')

            # Stream results
            im0 = annotator.result()
            if show_vid:
                global count, count_reg, count_tiff
                color=(0,255,0)
                #start_point = (w//3, h//4)   #Vertical Line
                #end_point = (w//3, (3*h)//4)
                start_point = (0, h//2)   #horizontal Line
                end_point = (w, h//2)
                
                #start_point = (0, h//2)    #Horizontal line To see the counting region when needed
                #end_point = (w, h//2)
                
                cv2.line(im0, start_point, end_point, color, thickness=2)
                thickness = 2
                org = (70, 110)
                font = cv2.FONT_HERSHEY_SIMPLEX
                fontScale = 1
                cv2.putText(im0,"EnergyPlus_Regular: "+str(count_reg), org, font,
                   fontScale, color, thickness, cv2.LINE_AA)            
                
                org = (70, 160)
                font = cv2.FONT_HERSHEY_SIMPLEX
                fontScale = 1
                cv2.putText(im0, "EnergyPlus_Tiffin: "+str(count_tiff), org, font, 
                   fontScale, color, thickness, cv2.LINE_AA)
                cv2.imshow(str(p), im0)            
                if cv2.waitKey(1) == ord('q'):  # q to quit
                    # # Workbook() takes one, non-optional, argument
                    # # which is the filename that we want to create.
                    # workbook = xlsxwriter.Workbook('Output_Data.xlsx')
                    
                    # # The workbook object is then used to add new
                    # # worksheet via the add_worksheet() method.
                    # worksheet = workbook.add_worksheet()
                    
                    # # Use the worksheet object to write
                    # # data via the write() method.
                    # worksheet.write('A1', "Product Name")
                    # worksheet.write('B1', "Total Counts")
                    # worksheet.write('A2', "EnergyPlus Regular = ")
                    # worksheet.write('B2', count_reg)
                    # worksheet.write('A3', "EnergyPlus Tiffin = ")
                    # worksheet.write('B3', count_tiff)
                    
                    # now = datetime.datetime.now()
                    # worksheet.write('A5', str(now))
                    
                    
                    
                    
                    # # Finally, close the Excel file
                    # # via the close() method.
                    # workbook.close()
                    
                    
                    # Check if the file exists
                    output_file_path = "Output_Data.xlsx"
                    

                    if os.path.exists(output_file_path):
                        # File exists, open it and update data
                        workbook = openpyxl.load_workbook(output_file_path)
                        worksheet = workbook.active
                        row_index = worksheet.max_row + 1  # Find the next available row

                        # Update your data here
                        count_reg = 10
                        count_tiff = 20

                        worksheet.cell(row=row_index, column=1, value="EnergyPlus Regular = ")
                        worksheet.cell(row=row_index, column=2, value=count_reg)
                        worksheet.cell(row=row_index + 1, column=1, value="EnergyPlus Tiffin = ")
                        worksheet.cell(row=row_index + 1, column=2, value=count_tiff)

                        now = datetime.datetime.now()
                        worksheet.cell(row=row_index + 3, column=1, value=str(now))

                    else:
                        # File doesn't exist, create it and add data
                        workbook = xlsxwriter.Workbook(output_file_path)
                        worksheet = workbook.add_worksheet()

                        worksheet.write('A1', "Product Name")
                        worksheet.write('B1', "Total Counts")
                        worksheet.write('A2', "EnergyPlus Regular = ")
                        worksheet.write('B2', count_reg)
                        worksheet.write('A3', "EnergyPlus Tiffin = ")
                        worksheet.write('B3', count_tiff)

                        now = datetime.datetime.now()
                        worksheet.write('A5', str(now))

                    # Finally, close the Excel file
                    workbook.close()                    
                    
                    
                    raise StopIteration

            # Save results (image with detections)
            if save_vid:
                if vid_path != save_path:  # new video
                    vid_path = save_path
                    if isinstance(vid_writer, cv2.VideoWriter):
                        vid_writer.release()  # release previous video writer
                    if vid_cap:  # video
                        fps = vid_cap.get(cv2.CAP_PROP_FPS)
                        w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                        h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    else:  # stream
                        fps, w, h = 30, im0.shape[1], im0.shape[0]

                    vid_writer = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (w, h))
                vid_writer.write(im0)

    # Print results
    t = tuple(x / seen * 1E3 for x in dt)  # speeds per image
    LOGGER.info(f'Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS, %.1fms deep sort update \
        per image at shape {(1, 3, *imgsz)}' % t)

    if save_txt or save_vid:
        print('Results saved to %s' % save_path)
        if platform == 'darwin':  # MacOS
            os.system('open ' + save_path)
"""
def count_obj(box,w,h,id,cls):
    #global count_Regular,data_Regular,count_Tiffin,data_Tiffin
    global count,data
    center_coordinates = (int(box[0]+(box[2]-box[0])/2) , int(box[1]+(box[3]-box[1])/2))
    if int(box[1]+(box[3]-box[1])/2) > (h -350):
      if id not in data:
       count += 1
       data.append(id)
"""

#Count Classes   
def count_obj(box,w,h,id,cls):
       global count_reg, data_reg, count_tiff , data_tiff
       center_coordinates = (int(box[0]+(box[2]-box[0])/2) , int(box[1]+(box[3]-box[1])/2))      
      
       if cls == 0:                    
            if int(box[1]+(box[3]-box[1])/2) >  (h//2):
                if id not in data_reg:
                    count_reg = count_reg+1
                    data_reg.append(id)
                   
       elif  cls == 1:
            if int(box[1]+(box[3]-box[1])/2) > (h//2):
                if id not in data_tiff:
                    count_tiff = count_tiff+1
                    data_tiff.append(id)



       # if  id not in data:
        #    count += 1
         #   data.append(id)




                

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--yolo_model', nargs='+', type=str, default="D:\Tracker\Test1\Yolov5_DeepSort_Pytorch\Custom_Energy_best.pt", help='model.pt path(s)')
    parser.add_argument('--deep_sort_model', type=str, default='osnet_x0_25')
    parser.add_argument('--source', type=str, default='D:\Training_Data\22.03.22\ch01_20220322140000.mp4', help='source')  # file/folder, 0 for webcam
    parser.add_argument('--output', type=str, default='inference/output', help='output folder')  # output folder
    parser.add_argument('--imgsz', '--img', '--img-size', nargs='+', type=int, default=[640], help='inference size h,w')
    parser.add_argument('--conf-thres', type=float, default=0.5, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.5, help='IOU threshold for NMS')
    parser.add_argument('--fourcc', type=str, default='mp4v', help='output video codec (verify ffmpeg support)')
    parser.add_argument('--device', default="0", help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--show-vid', action='store_false', help='display tracking video results')
    parser.add_argument('--save-vid', action='store_true', help='save video tracking results')
    parser.add_argument('--save-txt', action='store_true', help='save MOT compliant results to *.txt')
    # class 0 is person, 1 is bycicle, 2 is car... 79 is oven
    #parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --class 0, or --class 16 17')
    parser.add_argument('--classes', nargs='+', default=[0,1], type=int, help='filter by class')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--evaluate', action='store_true', help='augmented inference')
    parser.add_argument("--config_deepsort", type=str, default="deep_sort/configs/deep_sort.yaml")
    parser.add_argument("--half", action="store_true", help="use FP16 half-precision inference")
    parser.add_argument('--visualize', action='store_true', help='visualize features')
    parser.add_argument('--max-det', type=int, default=1000, help='maximum detection per image')
    parser.add_argument('--dnn', action='store_true', help='use OpenCV DNN for ONNX inference')
    parser.add_argument('--project', default=ROOT / 'runs/track', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    opt = parser.parse_args()
    opt.imgsz *= 2 if len(opt.imgsz) == 1 else 1  # expand

    with torch.no_grad():
        detect(opt)
